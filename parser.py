import os
import openpyxl
import statistics
import matplotlib.pyplot as plt
from utils import valid_file_name

column_headers = {
    'arg_1': 'Voltage',
    'arg_2': 'Qh',
    'arg_3': 'consumption',
    'arg_4': 'Th',
    'Qiv': 'Qiv',
    'Tin': 'Tin',
    'Tout': 'Tout'
}


def get_sheet(file_name: str):
    full_path = os.path.join('files', file_name)
    book = openpyxl.load_workbook(full_path, read_only=True)
    sheet = book.active
    return sheet


def convert_consumption(sheet, column_headers):
    consumption_col = get_column_by_header(sheet, column_headers['arg_3'])
    if consumption_col is None:
        return None
    consumption = sheet.cell(row=2, column=consumption_col).value
    if consumption is not None:
        return round((consumption * 60_000_000) / 100) * 100
    return None


def get_column_by_header(sheet, header_name: str) -> int | None:
    for col_num, cell in enumerate(sheet[1], start=1):
        if cell.value == header_name:
            return col_num
    return None


def get_unique_voltage_value(sheet, voltage_idx):
    arg_values = set()
    for row in sheet.iter_rows(values_only=True):
        if len(row) > voltage_idx:
            value = row[voltage_idx]
            if value is not None:
                try:
                    arg_values.add(float(value))
                except (ValueError, TypeError):
                    pass
    return sorted(arg_values)


def get_temperature_by_voltage(sheet, temp_idx, voltage_idx):
    temp_by_voltage = {}
    for row in sheet.iter_rows(values_only=True):
        if len(row) > max(voltage_idx, temp_idx):
            voltage = row[voltage_idx]
            temperature = row[temp_idx]
            if voltage is not None and temperature is not None:
                try:
                    temp_float = float(temperature)
                except (ValueError, TypeError):
                    continue
                if voltage not in temp_by_voltage:
                    temp_by_voltage[voltage] = []
                temp_by_voltage[voltage].append(temp_float)
    return temp_by_voltage

# ========== ОСНОВНАЯ ПРОГРАММА ==========


if __name__ == '__main__':
    sheet_list = valid_file_name()
    for filename in sheet_list:
        sheet = get_sheet(filename)

        voltage_column_number = get_column_by_header(sheet, column_headers['arg_1'])
        qh_column_number = get_column_by_header(sheet, column_headers['arg_2'])
        temp_column_number = get_column_by_header(sheet, column_headers['arg_4'])

        if None in [voltage_column_number, qh_column_number]:
            print("❌ Ошибка: не найдены колонки Voltage или Qh!")
            exit()

        voltage_idx = voltage_column_number - 1
        qh_idx = qh_column_number - 1
        temp_idx = temp_column_number - 1 if temp_column_number else None

        unique_voltages = get_unique_voltage_value(sheet, voltage_idx)

        consumption_value = convert_consumption(sheet, column_headers)

        qh_by_voltage = {voltage: [] for voltage in unique_voltages}

        for row in sheet.iter_rows(values_only=True):
            if len(row) > max(voltage_idx, qh_idx):
                voltage = row[voltage_idx]
                qh = row[qh_idx]
                if voltage in qh_by_voltage and qh is not None:
                    qh_by_voltage[voltage].append(qh)

        # Сбор данных температур
        if temp_idx is not None:
            temp_by_voltage = get_temperature_by_voltage(sheet, temp_idx, voltage_idx)
        else:
            temp_by_voltage = {}

        # Вывод
        print(f"\n📊 Статистика Qh и Th по напряжениям при расходе = {consumption_value} мл/мин:")
        print("="*80)
        print(f"{'Voltage':<12} {'Медиана Qh':<15} {'Медиана Th':<15}")
        print("="*80)

        all_temperatures = []
        all_qh = []

        for voltage in sorted(unique_voltages):
            qh_values = qh_by_voltage.get(voltage, [])
            temp_values = temp_by_voltage.get(voltage, [])

            all_temperatures.extend(temp_values)
            all_qh.extend(qh_values)

            qh_median = statistics.median(qh_values) if qh_values else None
            temp_median = statistics.median(temp_values) if temp_values else None

            if qh_median and temp_median:
                print(f"{voltage:<12.1f} {qh_median:<15.4f}  {temp_median:<15.2f}")

        # Построение графика
        plt.figure(figsize=(10, 6))
        plt.scatter(all_qh, all_temperatures, alpha=0.5, s=10, color='blue', label='Экспериментальные точки')

        # Настройки графика
        plt.xlabel('Qh (тепло, Вт)', fontsize=12)
        plt.ylabel('Температура Th (°C)', fontsize=12)
        plt.title(f'Зависимость температуры Th от Qh при расходе = {consumption_value} мл/мин', fontsize=14)
        plt.grid(True, alpha=0.3)
        plt.legend()

        # Показать график
        plt.show()
