import openpyxl
import statistics
import matplotlib.pyplot as plt

from utils import valid_file_name

column_headers = {
    'arg_1': 'Voltage',
    'arg_2': 'Qh',
    'arg_3': 'consumption',
    'arg_4': 'Th'
}

def get_sheet(file_name: str):
    book = openpyxl.load_workbook(file_name, read_only=True)
    sheet = book.active
    return sheet

def convert_consumption(sheet, column_headers):
    consumption_col = get_column_by_header(sheet, column_headers['arg_3'])
    if consumption_col is None:
        return None
    consumption = sheet.cell(row=2, column=consumption_col).value
    if consumption is not None:
        return round((consumption * 60_000_000) / 100) * 100
    return None

def get_column_by_header(sheet, header_name: str) -> int | None:
    for col_num, cell in enumerate(sheet[1], start=1):
        if cell.value == header_name:
            return col_num
    return None

def get_unique_voltage_value(sheet, voltage_idx):
    arg_values = set()
    for row in sheet.iter_rows(values_only=True):
        if len(row) > voltage_idx:
            value = row[voltage_idx]
            if value is not None:
                try:
                    arg_values.add(float(value))
                except (ValueError, TypeError):
                    pass
    return sorted(arg_values)

def get_temperature_by_voltage(sheet, temp_idx, voltage_idx):
    temp_by_voltage = {}
    for row in sheet.iter_rows(values_only=True):
        if len(row) > max(voltage_idx, temp_idx):
            voltage = row[voltage_idx]
            temperature = row[temp_idx]
            if voltage is not None and temperature is not None:
                try:
                    temp_float = float(temperature)
                except (ValueError, TypeError):
                    continue
                if voltage not in temp_by_voltage:
                    temp_by_voltage[voltage] = []
                temp_by_voltage[voltage].append(temp_float)
    return temp_by_voltage

# ========== ОСНОВНАЯ ПРОГРАММА ==========

if __name__ == '__main__':
    sheet_list = valid_file_name()
    
    if not sheet_list:
        print("❌ Не найдено Excel файлов!")
        exit()
    
    # Словари для сбора данных со всех файлов
    all_data_by_consumption = {}  # {расход: {'voltage': [], 'qh_median': [], 'temp_median': []}}
    
    for filename in sheet_list:
        print(f"\n📊 Обработка файла: {filename}")
        sheet = get_sheet(filename)

        voltage_column_number = get_column_by_header(sheet, column_headers['arg_1'])
        qh_column_number = get_column_by_header(sheet, column_headers['arg_2'])
        temp_column_number = get_column_by_header(sheet, column_headers['arg_4'])

        # Проверка
        if None in [voltage_column_number, qh_column_number]:
            print("❌ Ошибка: не найдены колонки Voltage или Qh! Пропускаем файл.")
            continue

        voltage_idx = voltage_column_number - 1
        qh_idx = qh_column_number - 1
        temp_idx = temp_column_number - 1 if temp_column_number else None

        unique_voltages = get_unique_voltage_value(sheet, voltage_idx)
        consumption_value = convert_consumption(sheet, column_headers)

        qh_by_voltage = {voltage: [] for voltage in unique_voltages}

        for row in sheet.iter_rows(values_only=True):
            if len(row) > max(voltage_idx, qh_idx):
                voltage = row[voltage_idx]
                qh = row[qh_idx]
                if voltage in qh_by_voltage and qh is not None:
                    qh_by_voltage[voltage].append(qh)

        # Сбор данных температур
        if temp_idx is not None:
            temp_by_voltage = get_temperature_by_voltage(sheet, temp_idx, voltage_idx)
        else:
            temp_by_voltage = {}

        # Вычисляем медианы для каждого напряжения
        voltages_sorted = sorted(unique_voltages)
        qh_medians = []
        temp_medians = []
        
        print(f"\n📊 Статистика для расхода {consumption_value} мл/мин:")
        print("="*80)
        print(f"{'Voltage':<12} {'Медиана Qh':<15} {'Медиана Th':<15}")
        print("="*80)
        
        for voltage in voltages_sorted:
            qh_values = qh_by_voltage.get(voltage, [])
            temp_values = temp_by_voltage.get(voltage, [])
            
            qh_median = statistics.median(qh_values) if qh_values else None
            temp_median = statistics.median(temp_values) if temp_values else None
            
            qh_medians.append(qh_median)
            temp_medians.append(temp_median)
            
            if qh_median and temp_median:
                print(f"{voltage:<12.1f} {qh_median:<15.4f}  {temp_median:<15.2f}")
            elif qh_median:
                print(f"{voltage:<12.1f} {qh_median:<15.4f}  {'N/A':<15}")
            elif temp_median:
                print(f"{voltage:<12.1f} {'N/A':<15}  {temp_median:<15.2f}")
            else:
                print(f"{voltage:<12.1f} {'Нет данных':<15}  {'Нет данных':<15}")
        
        # Сохраняем данные для построения графика (только где есть оба значения)
        valid_data = [(v, q, t) for v, q, t in zip(voltages_sorted, qh_medians, temp_medians) 
                     if q is not None and t is not None]
        
        if valid_data:
            voltages_clean, qh_clean, temp_clean = zip(*valid_data)
            
            if consumption_value not in all_data_by_consumption:
                all_data_by_consumption[consumption_value] = {'voltage': [], 'qh': [], 'temp': []}
            
            all_data_by_consumption[consumption_value]['voltage'].extend(voltages_clean)
            all_data_by_consumption[consumption_value]['qh'].extend(qh_clean)
            all_data_by_consumption[consumption_value]['temp'].extend(temp_clean)
    
    # ========== ПОСТРОЕНИЕ ГРАФИКА В ВИДЕ ЛИНИЙ ==========
    if all_data_by_consumption:
        plt.figure(figsize=(12, 8))
        
        # Цвета для разных расходов
        colors = ['blue', 'red', 'green', 'orange', 'purple', 'brown', 'pink', 'gray', 'olive', 'cyan']
        
        # График 1: Зависимость Qh от Voltage
        plt.subplot(1, 2, 1)
        for i, (consumption, data) in enumerate(sorted(all_data_by_consumption.items())):
            color = colors[i % len(colors)]
            
            # Сортируем точки по напряжению
            points = sorted(zip(data['voltage'], data['qh']))
            voltages_sorted, qh_sorted = zip(*points)
            
            plt.plot(voltages_sorted, qh_sorted, 
                    color=color, linewidth=2, marker='o', markersize=6,
                    label=f'Расход = {consumption} мл/мин')
        
        plt.xlabel('Voltage', fontsize=12)
        plt.ylabel('Qh (тепло, Вт)', fontsize=12)
        plt.title('Зависимость Qh от напряжения', fontsize=14)
        plt.grid(True, alpha=0.3)
        plt.legend(loc='best', fontsize=10)
        
        # График 2: Зависимость Th от Voltage
        plt.subplot(1, 2, 2)
        for i, (consumption, data) in enumerate(sorted(all_data_by_consumption.items())):
            color = colors[i % len(colors)]
            
            # Сортируем точки по напряжению
            points = sorted(zip(data['voltage'], data['temp']))
            voltages_sorted, temp_sorted = zip(*points)
            
            plt.plot(voltages_sorted, temp_sorted, 
                    color=color, linewidth=2, marker='s', markersize=6,
                    label=f'Расход = {consumption} мл/мин')
        
        plt.xlabel('Voltage', fontsize=12)
        plt.ylabel('Температура Th (°C)', fontsize=12)
        plt.title('Зависимость температуры от напряжения', fontsize=14)
        plt.grid(True, alpha=0.3)
        plt.legend(loc='best', fontsize=10)
        
        plt.tight_layout()
        plt.show()
        
        # График 3: Зависимость Th от Qh (линии)
        plt.figure(figsize=(10, 8))
        for i, (consumption, data) in enumerate(sorted(all_data_by_consumption.items())):
            color = colors[i % len(colors)]
            
            # Сортируем точки по Qh
            points = sorted(zip(data['qh'], data['temp']))
            qh_sorted, temp_sorted = zip(*points)
            
            plt.plot(qh_sorted, temp_sorted, 
                    color=color, linewidth=2, marker='o', markersize=6,
                    label=f'Расход = {consumption} мл/мин')
        
        plt.xlabel('Qh (тепло, Вт)', fontsize=12)
        plt.ylabel('Температура Th (°C)', fontsize=12)
        plt.title('Зависимость температуры Th от Qh для разных расходов', fontsize=14)
        plt.grid(True, alpha=0.3)
        plt.legend(loc='best', fontsize=10)
        plt.tight_layout()
        plt.show()
        
        # Вывод общей статистики
        print("\n" + "="*80)
        print("📊 ОБЩАЯ СТАТИСТИКА ПО ВСЕМ ФАЙЛАМ:")
        print("="*80)
        for consumption, data in sorted(all_data_by_consumption.items()):
            qh_median = statistics.median(data['qh']) if data['qh'] else None
            temp_median = statistics.median(data['temp']) if data['temp'] else None
            print(f"Расход {consumption} мл/мин: "
                  f"Медиана Qh = {qh_median:.4f} Вт, "
                  f"Медиана Th = {temp_median:.2f} °C, "
                  f"Точек: {len(data['qh'])}")
    else:
        print("❌ Нет данных для построения графика!")